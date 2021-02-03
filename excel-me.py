import sys, os
from PIL import Image
import numpy as np
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter


def main(arg):

    path = arg[0]   # store filepath

    image = read_image(path)
    # Resize the image if there is --resize flag
    if "--resize" is in arg:
        image = resize_image(image, 360, 640)   # rezise into 360p
    list, width, height = image_to_list(image)
    #image.show()

    filename = get_filename(path)
    # Create an workbook instance
    wb = Workbook()
    # Select the active worksheet
    sheet = wb.active

    # Apend the pixel values to worksheet
    for row in list:
        sheet.append(row)

    # Make conditional formatting rules
    color_scale_rule_red = ColorScaleRule(start_type="num",
                                      start_value=0,
                                      start_color="000000",
                                      end_type="num",
                                      end_value=255,
                                      end_color="FF0000")
    color_scale_rule_green = ColorScaleRule(start_type="num",
                                      start_value=0,
                                      start_color="000000",
                                      end_type="num",
                                      end_value=255,
                                      end_color="00FF00")
    color_scale_rule_blue = ColorScaleRule(start_type="num",
                                      start_value=0,
                                      start_color="000000",
                                      end_type="num",
                                      end_value=255,
                                      end_color="0000FF")

    # Formula: col_width = (col_height * 0.175) / 3
    col_height = 12.75
    col_width = (col_height * 0.175) / 3

    for i in range(1, width + 1):
        # Convert column number to column letter
        col = get_column_letter(i)
        # Make the range string
        string = col + "1" + ":" + col + str(height)
        # Set column height, width for the cells
        sheet.column_dimensions[col].height = col_height
        sheet.column_dimensions[col].width = col_width
        # Add conditional formatting based on rules
        if i % 3 == 1:
            sheet.conditional_formatting.add(string, color_scale_rule_red)
        elif i % 3 == 2:
            sheet.conditional_formatting.add(string, color_scale_rule_green)
        elif i % 3 == 0:
            sheet.conditional_formatting.add(string, color_scale_rule_blue)

    # Create the excel file
    wb.save(filename=filename)


def image_to_list(image):
    """Convert image to tabular form and retun the list, width and height of the table."""
    arr = np.array(image)
    h, w, p = arr.shape[0], arr.shape[1], arr.shape[2]
    # Reshape to include RGB pixels in the column dimension
    arr = arr.reshape((h, w*p))
    list = arr.tolist()
    width = (w*p)
    height = h
    return list, width, height


def resize_image(image, height, width):
    """Resize the image (to work with less values)."""
    return image.resize((height, width))


def read_image(path):
    """Read an image from a file path"""
    try:
        image = Image.open(path)
        return image
    except Exception as e:
        print(e)


def get_filename(path):
    """Parse filepath, remove old file and retuen the filename."""
    # Take the filename part of the path (in lowercase)
    name = os.path.basename(path).lower()
    # Remove file extension
    name = name.replace(".jpeg", "").replace(".jpg", "").replace(".png", "")
    name += ".xlsx"
    # Remove old *.xlsx file
    if os.path.exists(name):
        os.remove(name)
    return name


if __name__ == "__main__":
    main(sys.argv[1:])
